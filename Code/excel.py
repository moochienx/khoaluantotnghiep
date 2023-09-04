from openpyxl import *
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
import re

class Excel:
    def open_excel_file(self, file_name):
        try:
            # Đọc tập tin excel và lưu vào DataFrame
            self.df = pd.read_excel(file_name)
        except Exception as e:
            return False

    def get_excel_file(self, file_name):
        self.open_excel_file(file_name)
        try:
            # Lấy ra số lượng hàng trong DataFrame
            num_rows = self.df.shape[0]
            # Xác định loại tệp danh sách lớp hoặc thi
            data_c = self.df.iloc[5, 4]
            data_c = data_c.find("thi")
            if data_c != -1:
                # Lấy ra dòng chưa số lượng sinh viên
                data = self.df.iloc[num_rows - 9, 0]
            elif data_c == -1:
                data = self.df.iloc[num_rows - 3, 0]
            else:
                return False

            # Sử dụng biểu thức chính quy để tìm số lượng sinh viên
            numbers = re.findall(r'\d+', data)

            # Ép kiểu số lượng sinh viên từ list sang int
            item = numbers[0]
            item = int(item)

            # Đọc tệp Excel và lưu vào một DataFrame
            df2 = pd.read_excel(file_name, usecols='B:F', skiprows=9, nrows=item)

            # Lấy dữ liệu trong DataFrame
            data2 = df2.iloc[:, :]
            return data2
        except Exception as e:
            return False
    def write_excel_file(self, file_name,data_checked):
        self.open_excel_file(file_name)
        # Lấy ra số lượng hàng trong DataFrame
        num_rows = self.df.shape[0]
        # Xác định loại tệp danh sách lớp hoặc thi
        data_c = self.df.iloc[5, 4]
        data_c2 = data_c.find("thi")
        if data_c2 != -1:
            # Lấy ra dòng chưa số lượng sinh viên
            data = self.df.iloc[num_rows - 9, 0]
            split_data = data_c.split(":")
            result = split_data[1].strip()
            result = result.split("/")
            result = result[0]
        elif data_c2 == -1:
            data = self.df.iloc[num_rows - 3, 0]
            split_data = data_c.split(":")
            result = split_data[1].strip()
        else:
            return False

        # Sử dụng biểu thức chính quy để tìm số lượng sinh viên
        numbers = re.findall(r'\d+', data)

        # Ép kiểu số lượng sinh viên từ list sang int
        item = numbers[0]
        item = int(item)

        # Lấy ra ngày, tháng, năm của ngày hiện tại
        now = datetime.now()
        day = now.day
        month = now.month
        year = now.year

        # Gán ngày, tháng, năm vào một biến
        today = f"{day}/{month}/{year}"

        # Tạo border cho cell
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        try:
            # Mở tập tin Excel
            wb = load_workbook(file_name)

            # Lấy sheet đầu tiên trong Workbook
            ws = wb.active

            # Khởi tạo biến chứa vị trí hàng cột theo file excel
            rw = 10
            col = 10

            # Vòng lặp xác định, ghi dữ liệu cột điểm danh
            while True:
                # Xác định vị trí của ô ghi ngày điểm danh
                cell = ws.cell(row=rw, column=col)
                # Kiểm tra xem ô (x, y) có tồn tại dữ liệu hay không
                if cell.value is None:
                    # Ghi dữ liệu vào ô (x, y)
                    cell.value = today
                    cell.border = border
                    for x in range(1, item + 1):
                        rw = rw + 1
                        value_cell = ws.cell(row=rw, column=col)
                        value_cell.value = (data_checked[x - 1])
                        value_cell.border = border
                        # Tự động lấy tên cột
                        column_letter = get_column_letter(col)
                        # Lấy ra kích thước lớn nhất giá trị trong cột
                        len_cell = len(cell.value)
                    break
                else:
                    col = col + 1
            # Căn chỉnh kích thước cột phù hợp với độ dài của dữ liệu
            ws.column_dimensions[column_letter].width = len_cell
            # Lưu lại tập tin Excel
            wb.save(file_name)
            return result
        except Exception as e:
            return False